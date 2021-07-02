from typing import List
from io import BytesIO

from openpyxl import load_workbook
from requests.models import Response

from .base_parser import BaseParser
from .data_types import RcBin

import sys
from openpyxl.cell.cell import MergedCell


class XlsxParser(BaseParser):
    def _cell_value(self,sheet,row,column):
        cell = sheet.cell(row,column)
        if not isinstance(cell, MergedCell):
          return cell.value

      # "Oh no, the cell is merged!"
        for range in sheet.merged_cells.ranges:
          if sheet.cell(row,column).coordinate in range:
            return range.start_cell.value

        return''

    def _get_list_from_resource(self, resource: Response) -> list:
        wb = load_workbook(BytesIO(resource.content))
        sheet = wb.active

        rc_bins = []

        for row in range(1,sheet.max_row):
          rows = []
          for column in range(1,sheet.max_column):
            self._cell_value(sheet, row, column)
            rows.append(column)
        rc_bins.append(rows)

        return rc_bins
